"""
/get_stats_excel - выгрузка личной статистики прокачки в Excel.

Данные берутся из отдельной таблицы daily_stats (см. bot/database.py), а не из
users/challenges - чтобы не раздувать основные таблицы: там хранится всего
одна строка на пользователя на календарный день с итоговыми числами по
каждой дисциплине. Строка пишется при закрытии дня (завершено / сдался /
просрочено) через Database.record_daily_stats.

Файл собирается на лету через openpyxl и отправляется как документ - на
диске ничего не остаётся дольше одного запроса.
"""
import io

from aiogram import Bot, Router
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from bot.database import db
from bot.utils import replace_command_message, send_command_message, track_command_message

router = Router(name="stats")

# Порядок и подписи колонок ровно как в запросе: Дата | Отжимания | Пресс |
# Приседания | Партии в шахматы | Прочитано страниц книги.
_COLUMNS = [
    ("date", "Дата"),
    ("pushups", "Отжимания"),
    ("abs", "Пресс"),
    ("squats", "Приседания"),
    ("chess", "Партии в шахматы"),
    ("reading", "Прочитано страниц книги"),
]


def _build_workbook(rows: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    header_font = Font(bold=True)
    for col_idx, (_, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])

    ws.freeze_panes = "A2"
    for col_idx, (key, label) in enumerate(_COLUMNS, start=1):
        width = max(len(label), 10 if key != "date" else 12) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.message(Command("get_stats_excel"))
async def cmd_get_stats_excel(message: Message, bot: Bot):
    user = await db.get_user(message.from_user.id)
    if not user or user["reg_state"] != "done":
        await send_command_message(bot, message.chat.id, "get_stats_excel", "Сначала пройди регистрацию: /start")
        return

    rows = await db.get_stats_rows(message.from_user.id)
    if not rows:
        await send_command_message(
            bot, message.chat.id, "get_stats_excel",
            "Пока нет ни одного завершённого дня - статистика появится после первого закрытого испытания.",
        )
        return

    buffer = _build_workbook(rows)
    file = BufferedInputFile(buffer.read(), filename="solo_leveling_stats.xlsx")

    # Документ отправляется отдельным вызовом (answer_document), поэтому
    # используем replace/track напрямую вместо send_command_message - но то же
    # правило "одно сообщение на команду" сохраняется: старый файл/ошибка этой
    # же команды убирается перед отправкой нового.
    await replace_command_message(bot, message.chat.id, "get_stats_excel")
    sent = await message.answer_document(file, caption="📊 Твоя статистика прокачки по дням.")
    track_command_message(bot, message.chat.id, "get_stats_excel", sent.message_id)
